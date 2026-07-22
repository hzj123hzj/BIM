#!/usr/bin/env python3
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import date

# Create workbook for 张三
wb1 = openpyxl.Workbook()
ws1 = wb1.active
ws1.title = "健康记录"

# Headers
headers = ["用户名", "记录日期", "体重(kg)", "体脂率(%)", "水分率(%)", "蛋白质率(%)", 
           "肌肉率(%)", "内脏脂肪", "骨骼肌(kg)", "骨量(kg)", "腰围(cm)"]
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")

for col, header in enumerate(headers, 1):
    cell = ws1.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

# Data for 张三
data_zs = [
    ["张三", date(2025,1,15), 72.5, 22.5, 58.3, 18.2, 42.1, 7, 38.5, 2.8, 82.5],
    ["张三", date(2025,2,15), 71.8, 21.8, 58.8, 18.5, 42.8, 6, 39.0, 2.8, 81.2],
    ["张三", date(2025,3,15), 71.2, 21.2, 59.2, 18.8, 43.2, 6, 39.5, 2.9, 80.5],
    ["张三", date(2025,4,15), 70.5, 20.5, 59.8, 19.0, 43.8, 5, 40.0, 2.9, 79.8],
    ["张三", date(2025,5,15), 69.8, 19.8, 60.2, 19.2, 44.2, 5, 40.5, 3.0, 79.0],
    ["张三", date(2025,6,15), 69.2, 19.2, 60.5, 19.5, 44.8, 4, 41.0, 3.0, 78.5],
]

for row_idx, row_data in enumerate(data_zs, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws1.cell(row=row_idx, column=col_idx, value=val)
        if col_idx == 2:
            cell.number_format = 'YYYY-MM-DD'
        cell.alignment = Alignment(horizontal="center")

# Auto-fit columns
for col in ws1.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws1.column_dimensions[col[0].column_letter].width = max_length + 4

wb1.save("/mnt/d/表格/病人表格/测试医疗中心_病人健康数据_张三.xlsx")
print("Created 张三.xlsx")

# Create workbook for 王林
wb2 = openpyxl.Workbook()
ws2 = wb2.active
ws2.title = "健康记录"

for col, header in enumerate(headers, 1):
    cell = ws2.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

data_wl = [
    ["王林", date(2025,1,15), 85.0, 28.5, 52.3, 16.2, 38.1, 12, 35.5, 2.5, 95.5],
    ["王林", date(2025,2,15), 84.2, 27.8, 53.0, 16.5, 38.8, 11, 36.0, 2.5, 94.2],
    ["王林", date(2025,3,15), 83.5, 27.2, 53.5, 16.8, 39.2, 11, 36.5, 2.6, 93.5],
    ["王林", date(2025,4,15), 82.8, 26.5, 54.0, 17.0, 39.8, 10, 37.0, 2.6, 92.8],
    ["王林", date(2025,5,15), 82.0, 25.8, 54.5, 17.2, 40.2, 10, 37.5, 2.7, 92.0],
    ["王林", date(2025,6,15), 81.2, 25.2, 55.0, 17.5, 40.8, 9, 38.0, 2.7, 91.2],
]

for row_idx, row_data in enumerate(data_wl, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws2.cell(row=row_idx, column=col_idx, value=val)
        if col_idx == 2:
            cell.number_format = 'YYYY-MM-DD'
        cell.alignment = Alignment(horizontal="center")

for col in ws2.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws2.column_dimensions[col[0].column_letter].width = max_length + 4

wb2.save("/mnt/d/表格/病人表格/测试医疗中心_病人健康数据_王林.xlsx")
print("Created 王林.xlsx")

# Create combined workbook
wb3 = openpyxl.Workbook()
ws3 = wb3.active
ws3.title = "批量导入"

for col, header in enumerate(headers, 1):
    cell = ws3.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

all_data = data_zs + data_wl
for row_idx, row_data in enumerate(all_data, 2):
    for col_idx, val in enumerate(row_data, 1):
        cell = ws3.cell(row=row_idx, column=col_idx, value=val)
        if col_idx == 2:
            cell.number_format = 'YYYY-MM-DD'
        cell.alignment = Alignment(horizontal="center")

for col in ws3.columns:
    max_length = max(len(str(cell.value)) for cell in col if cell.value)
    ws3.column_dimensions[col[0].column_letter].width = max_length + 4

wb3.save("/mnt/d/表格/病人表格/测试医疗中心_病人健康数据_批量导入.xlsx")
print("Created 批量导入.xlsx")

print("\nAll Excel files created successfully!")